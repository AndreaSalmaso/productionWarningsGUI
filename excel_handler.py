from openpyxl import *
from openpyxl.styles import Alignment
from datetime import datetime, date
from tkinter import *
import pandas as pd


def open_given_wb(filePath):
    global sheet 
    
    workbook = filePath
    w = load_workbook(workbook)
    sheet = w['Segnalazioni']


def add_data_to_sheet(filePath, report_data):

    global sheet
    workbook = filePath
    w = load_workbook(workbook)
    sheet = w['Segnalazioni']
   
    reporting_date = date.today().strftime("%d-%b-%Y")
    report_data.insert(5, reporting_date)
    reporting_hour = datetime.now().strftime("%H:%M")
    report_data.insert(6, reporting_hour)

    r = len(list(sheet.rows)) + 1
    for c, val in enumerate(report_data):
        sheet.cell(row=r, column=c+1).value = val
        if c==4:
            sheet.cell(row=r, column=c+1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        else:
            sheet.cell(row=r, column=c+1).alignment = Alignment(horizontal='center', vertical='center')

    w.save(workbook)


def get_item_codes(filePath):
    
    workbook = filePath
    w = load_workbook(workbook)
    sheet = w['Articoli']

    items = []
    i=2
    while sheet.cell(row=i, column=1).value is not None:
        items.append(sheet.cell(row=i, column=1).value)
        i=i+1

    return items


    # open_given_wb(filePath)
    # headers = [h.value for h in sheet[1]]
    # ordine, codice, ogg, seriale, probl, data, ora, op = [], [], [], [], [], [], [], []
    # first_empty_row = len(list(sheet.rows)) + 1
    # for row in sheet.iter_rows(min_col=1, max_col=8, min_row=first_empty_row - 3, max_row=first_empty_row - 1, values_only=True):
    #     for i, c in enumerate([ordine, codice, ogg, seriale, probl, data, ora, op]):
    #         c.append(row[i])

    # t = zip(ordine, codice, ogg, seriale, probl, data, ora, op)
    # return tabulate(t, headers=headers)


if __name__=='__main__':
    w = load_workbook('Dati/SegnalazioniProduzione.xlsx')
    sheet = w['Segnalazioni']

    print(len(list(sheet.rows)))
